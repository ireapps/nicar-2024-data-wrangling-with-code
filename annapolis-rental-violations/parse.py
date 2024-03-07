from openpyxl import load_workbook

import datetime
import csv


wb = load_workbook('2020-city-of-annapolis-rental-violations.xlsx')

ws = wb['Sheet1']

# list of headers for parsed file
headers = [
    'case_no',
    'date_opened',
    'subtype',
    'date_closed',
    'assigned_to',
    'site_address',
    'parcel_no',
    'owner',
    'status',
    'violation_description'
]

OUTFILE = 'data-parsed.csv'

data = []
d = {}

for key, *values in ws.iter_rows(min_row=14):
    row = [x.value for x in values]

    if all(x is None for x in row):
        continue

    if row[0]:
        if str(row[0]).startswith('RENT') and str(row[0]) != 'RENTAL':

            if d:
                data.append(d)

            d = dict.fromkeys(headers, '')
            d['case_no'] = row[0]
            d['date_opened'] = row[1].date()
            d['subtype'] = row[3]
            d['assigned_to'] = row[4]
            d['site_address'] = row[5]
            d['owner'] = row[6]
    else:
        if isinstance(row[1], datetime.datetime):
            d['date_closed'] = row[1].date()
            d['status'] = row[4]
            d['parcel_no'] = row[5]
        else:
            d['violation_description'] = row[2]

with open(OUTFILE, 'w') as out:
    writer = csv.DictWriter(out, fieldnames=headers)
    writer.writeheader()
    writer.writerows(data)

print(f'Wrote {OUTFILE}')
